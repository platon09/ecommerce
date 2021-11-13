from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from config.settings import DATA_DIR
from market.models import Category, Product

class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print(f'Start importing from excel {DATA_DIR}!')

        file_name = 'тауар.xlsx'
        wb = load_workbook(DATA_DIR+'/'+file_name)
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])

        number_row = sheet.max_row
        for cnt in range(2, number_row):
            prod_id = sheet.cell(row=cnt, column=1).value
            item = sheet.cell(row=cnt, column=2).value
            amount = sheet.cell(row=cnt, column=3).value
            price = sheet.cell(row=cnt, column=4).value

            if prod_id:
                product = Product()
                product.name = item
                product.category = category
                product.save()
            else:
                category = Category()
                category.name = item
                category.save()
